#!/usr/bin/env python3
"""Local LanceDB RAG index for Obsidian/text files."""

from __future__ import annotations

import argparse
import fnmatch
import hashlib
import io
import json
import os
import re
import shlex
import shutil
import subprocess
import sys
import tempfile
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable


PROJECT_ROOT = Path(__file__).resolve().parents[1]
ENV_FILE = Path(os.environ.get("OMLX_HOME", str(PROJECT_ROOT))) / ".env"
TEXT_EXTENSIONS = ".md,.txt,.rst,.csv,.tsv,.json,.yaml,.yml,.toml,.xml,.html"
SPREADSHEET_EXTENSIONS = ".xlsx,.xlsm,.xls,.xlsb,.ods"
PDF_EXTENSIONS = ".pdf"
IMAGE_EXTENSIONS = ".png,.jpg,.jpeg,.tif,.tiff"
DEFAULT_EXTENSIONS = f"{TEXT_EXTENSIONS},{SPREADSHEET_EXTENSIONS},{PDF_EXTENSIONS},{IMAGE_EXTENSIONS}"
DEFAULT_EXCLUDES = ".git/**,.obsidian/**,node_modules/**,.trash/**,*.env,*.key,*.pem"
TABLE_NAME = "chunks"
MANIFEST_NAME = "manifest.json"
TABLE_SCHEMA_VERSION = 3
SPREADSHEET_EXT_SET = {".xlsx", ".xlsm", ".xls", ".xlsb", ".ods"}
PDF_EXT_SET = {".pdf"}
IMAGE_EXT_SET = {".png", ".jpg", ".jpeg", ".tif", ".tiff"}


def read_env_file(path: Path = ENV_FILE) -> dict[str, str]:
    values: dict[str, str] = {}
    if not path.exists():
        return values
    for raw_line in path.read_text(encoding="utf-8", errors="replace").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip()
        if (value.startswith('"') and value.endswith('"')) or (
            value.startswith("'") and value.endswith("'")
        ):
            value = value[1:-1]
        values.setdefault(key, os.path.expandvars(value))
    return values


ENV = read_env_file()


def env(name: str, default: str = "") -> str:
    value = os.environ.get(name, ENV.get(name, default))
    return os.path.expandvars(value)


def env_bool(name: str, default: bool = False) -> bool:
    value = env(name, "1" if default else "0").strip().lower()
    return value in {"1", "true", "yes", "on"}


def env_int(name: str, default: int) -> int:
    try:
        return int(env(name, str(default)))
    except ValueError:
        return default


def expand_path(value: str) -> Path:
    return Path(os.path.expandvars(os.path.expanduser(value))).resolve()


def split_csv(value: str) -> list[str]:
    return [item.strip() for item in value.split(",") if item.strip()]


def normalize_exts(value: str) -> set[str]:
    exts: set[str] = set()
    for ext in split_csv(value):
        exts.add(ext if ext.startswith(".") else f".{ext}")
    return exts


def source_path() -> Path:
    raw = env("RAG_SOURCE_PATH", env("OBSIDIAN_SHARED_PATH", ""))
    if raw in {"", "${OBSIDIAN_SHARED_PATH}", "${OBSIDIAN_SHARED_PATH:-}"}:
        raw = env("OBSIDIAN_SHARED_PATH", "")
    if not raw:
        raise RuntimeError("RAG_SOURCE_PATH is unset; set OBSIDIAN_SHARED_PATH or RAG_SOURCE_PATH.")
    return expand_path(raw)


def index_path() -> Path:
    raw = env("RAG_INDEX_PATH", ".runtime/rag")
    path = expand_path(raw)
    if not path.is_absolute():
        path = (PROJECT_ROOT / raw).resolve()
    return path


def lancedb_path() -> Path:
    return index_path() / "lancedb"


def manifest_path() -> Path:
    return index_path() / MANIFEST_NAME


def load_manifest() -> dict[str, Any]:
    path = manifest_path()
    if not path.exists():
        return {"documents": {}}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {"documents": {}}


def save_manifest(manifest: dict[str, Any]) -> None:
    path = manifest_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(".tmp")
    tmp.write_text(json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")
    tmp.replace(path)


def relpath(path: Path, root: Path) -> str:
    return path.resolve().relative_to(root.resolve()).as_posix()


def is_excluded(relative: str, patterns: Iterable[str]) -> bool:
    parts = Path(relative).parts
    for pattern in patterns:
        pattern = pattern.strip()
        if not pattern:
            continue
        if fnmatch.fnmatch(relative, pattern) or fnmatch.fnmatch(Path(relative).name, pattern):
            return True
        if pattern.endswith("/**") and relative.startswith(pattern[:-3].rstrip("/") + "/"):
            return True
        if pattern.startswith(".") and pattern.endswith("/**"):
            dirname = pattern[:-3].rstrip("/")
            if dirname in parts:
                return True
        if pattern.endswith("/**"):
            dirname = pattern[:-3].rstrip("/")
            if "/" not in dirname and dirname in parts:
                return True
    return False


def discover_files(root: Path) -> list[Path]:
    extensions = normalize_exts(env("RAG_TEXT_EXTENSIONS", DEFAULT_EXTENSIONS))
    excludes = split_csv(env("RAG_EXCLUDE_GLOBS", DEFAULT_EXCLUDES))
    files: list[Path] = []
    for path in root.rglob("*"):
        if not path.is_file():
            continue
        relative = relpath(path, root)
        if is_excluded(relative, excludes):
            continue
        if path.suffix.lower() not in extensions:
            continue
        try:
            max_mb = env_int("RAG_MAX_FILE_MB", 10)
            if path.suffix.lower() in SPREADSHEET_EXT_SET:
                max_mb = env_int("RAG_SPREADSHEET_MAX_FILE_MB", env_int("RAG_DOCUMENT_MAX_FILE_MB", 50))
            elif path.suffix.lower() in PDF_EXT_SET | IMAGE_EXT_SET:
                max_mb = env_int("RAG_DOCUMENT_MAX_FILE_MB", 50)
            if path.stat().st_size > max_mb * 1024 * 1024:
                continue
        except OSError:
            continue
        files.append(path)
    return sorted(files)


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


FRONTMATTER_RE = re.compile(r"\A---\s*\n(.*?)\n---\s*\n", re.DOTALL)
HEADING_RE = re.compile(r"^(#{1,6})\s+(.+?)\s*$", re.MULTILINE)
TAG_RE = re.compile(r"(?<!\w)#([A-Za-z0-9_/-]+)")
WIKILINK_RE = re.compile(r"\[\[([^\]]+)\]\]")


def parse_frontmatter(text: str) -> tuple[dict[str, Any], str]:
    match = FRONTMATTER_RE.match(text)
    if not match:
        return {}, text
    raw = match.group(1)
    frontmatter: dict[str, Any] = {}
    for line in raw.splitlines():
        if ":" not in line or line.startswith((" ", "\t", "-")):
            continue
        key, value = line.split(":", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if value.startswith("[") and value.endswith("]"):
            value = [item.strip().strip('"').strip("'") for item in value[1:-1].split(",") if item.strip()]
        frontmatter[key] = value
    return frontmatter, text[match.end() :]


def extract_tags(text: str, frontmatter: dict[str, Any]) -> list[str]:
    tags = set(TAG_RE.findall(text))
    raw_tags = frontmatter.get("tags") or frontmatter.get("tag")
    if isinstance(raw_tags, str):
        tags.update(item.strip().lstrip("#") for item in re.split(r"[,\s]+", raw_tags) if item.strip())
    elif isinstance(raw_tags, list):
        tags.update(str(item).strip().lstrip("#") for item in raw_tags if str(item).strip())
    return sorted(tags)


def extract_links(text: str) -> list[str]:
    links = []
    for raw in WIKILINK_RE.findall(text):
        target = raw.split("|", 1)[0].split("#", 1)[0].strip()
        if target:
            links.append(target)
    return sorted(set(links))


def extract_title(path: Path, frontmatter: dict[str, Any], text: str) -> str:
    for key in ("title", "name"):
        value = frontmatter.get(key)
        if isinstance(value, str) and value.strip():
            return value.strip()
    match = HEADING_RE.search(text)
    if match:
        return match.group(2).strip()
    return path.stem


def tokenish_len(text: str) -> int:
    return max(1, len(text) // 4)


@dataclass
class Chunk:
    text: str
    heading: str
    source_type: str = "text"
    sheet: str = ""
    cell_range: str = ""
    row_start: int = 0
    row_end: int = 0
    column_headers: str = ""
    has_formulas: bool = False
    hidden_sheet: bool = False
    page: int = 0
    page_start: int = 0
    page_end: int = 0
    ocr_used: bool = False
    ocr_languages: str = ""
    extractor: str = "text"


def trim_to_chunks(blocks: list[tuple[str, str]], target_tokens: int, overlap_tokens: int) -> list[Chunk]:
    target_chars = max(400, target_tokens * 4)
    overlap_chars = max(0, overlap_tokens * 4)
    chunks: list[Chunk] = []
    current = ""
    current_heading = ""
    for heading, block in blocks:
        block = block.strip()
        if not block:
            continue
        if len(block) > target_chars:
            start = 0
            while start < len(block):
                part = block[start : start + target_chars].strip()
                if part:
                    chunks.append(Chunk(part, heading))
                if start + target_chars >= len(block):
                    break
                start = max(start + target_chars - overlap_chars, start + 1)
            current = ""
            current_heading = ""
            continue
        if current and len(current) + len(block) + 2 > target_chars:
            chunks.append(Chunk(current.strip(), current_heading))
            tail = current[-overlap_chars:].strip() if overlap_chars else ""
            current = f"{tail}\n\n{block}" if tail else block
            current_heading = heading
        else:
            current = f"{current}\n\n{block}" if current else block
            current_heading = current_heading or heading
    if current.strip():
        chunks.append(Chunk(current.strip(), current_heading))
    return chunks


def markdown_blocks(text: str) -> list[tuple[str, str]]:
    blocks: list[tuple[str, str]] = []
    current_heading = ""
    current_lines: list[str] = []
    for line in text.splitlines():
        heading = re.match(r"^(#{1,6})\s+(.+?)\s*$", line)
        if heading and current_lines:
            blocks.append((current_heading, "\n".join(current_lines)))
            current_lines = []
        if heading:
            current_heading = heading.group(2).strip()
        current_lines.append(line)
    if current_lines:
        blocks.append((current_heading, "\n".join(current_lines)))
    return blocks


def plain_blocks(text: str) -> list[tuple[str, str]]:
    parts = [part.strip() for part in re.split(r"\n\s*\n", text) if part.strip()]
    if not parts:
        parts = [line.strip() for line in text.splitlines() if line.strip()]
    return [("", part) for part in parts]


def stringify_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return f"{value:.12g}"
    return str(value).strip()


def excel_col_name(index: int) -> str:
    name = ""
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        name = chr(65 + remainder) + name
    return name or "A"


def trim_sheet_rows(rows: list[tuple[int, list[str]]]) -> tuple[list[tuple[int, list[str]]], int]:
    non_empty = [(row_num, row) for row_num, row in rows if any(cell for cell in row)]
    if not non_empty:
        return [], 1
    max_width = max(len(row) for _, row in non_empty)
    keep_columns = [
        col
        for col in range(max_width)
        if any(col < len(row) and row[col] for _, row in non_empty)
    ]
    if not keep_columns:
        return [], 1
    first_col = keep_columns[0] + 1
    trimmed = [
        (row_num, [row[col] if col < len(row) else "" for col in keep_columns])
        for row_num, row in non_empty
    ]
    return trimmed, first_col


def headers_from_rows(rows: list[tuple[int, list[str]]]) -> list[str]:
    if not rows:
        return []
    raw = rows[0][1]
    headers = []
    for index, value in enumerate(raw, start=1):
        value = value.strip()
        headers.append(value or f"Column {excel_col_name(index)}")
    return headers


def records_text(headers: list[str], rows: list[tuple[int, list[str]]]) -> str:
    lines = []
    data_rows = rows[1:] if headers and rows else rows
    for row_num, row in data_rows:
        pairs = []
        for index, value in enumerate(row):
            if not value:
                continue
            header = headers[index] if index < len(headers) else f"Column {excel_col_name(index + 1)}"
            pairs.append(f"{header}: {value}")
        if pairs:
            lines.append(f"Row {row_num}: " + "; ".join(pairs))
    return "\n".join(lines)


def spreadsheet_sheet_chunks(
    path: Path,
    sheet_name: str,
    rows: list[tuple[int, list[str]]],
    first_col: int,
    hidden: bool = False,
    formulas: list[str] | None = None,
    comments: list[str] | None = None,
) -> list[Chunk]:
    chunks: list[Chunk] = []
    formulas = formulas or []
    comments = comments or []
    headers = headers_from_rows(rows)
    row_count = max(0, len(rows) - (1 if headers else 0))
    col_count = max((len(row) for _, row in rows), default=0)
    last_row = rows[-1][0] if rows else 0
    last_col = first_col + col_count - 1 if col_count else first_col
    range_label = f"{excel_col_name(first_col)}{rows[0][0]}:{excel_col_name(last_col)}{last_row}" if rows else ""
    header_text = ", ".join(headers) if headers else "(no headers detected)"
    summary = (
        f"Spreadsheet file: {path.name}\n"
        f"Sheet: {sheet_name}\n"
        f"Visible: {'no' if hidden else 'yes'}\n"
        f"Range: {range_label or 'empty'}\n"
        f"Rows: {row_count}\n"
        f"Columns: {col_count}\n"
        f"Headers: {header_text}\n"
        f"Contains formulas: {'yes' if formulas else 'no'}\n"
        f"Contains comments: {'yes' if comments else 'no'}"
    )
    chunks.append(
        Chunk(
            summary,
            f"Sheet summary: {sheet_name}",
            source_type="spreadsheet",
            sheet=sheet_name,
            cell_range=range_label,
            row_start=rows[0][0] if rows else 0,
            row_end=last_row,
            column_headers=json.dumps(headers, ensure_ascii=False),
            has_formulas=bool(formulas),
            hidden_sheet=hidden,
        )
    )

    if not rows:
        return chunks

    max_rows_full = env_int("RAG_SPREADSHEET_MAX_ROWS_FULL", 5000)
    max_rows_per_chunk = max(1, env_int("RAG_SPREADSHEET_MAX_ROWS_PER_CHUNK", 50))
    row_limit = min(len(rows), max_rows_full + 1 if headers else max_rows_full)
    limited_rows = rows[:row_limit]
    data_start = 1 if headers else 0
    for start in range(data_start, len(limited_rows), max_rows_per_chunk):
        chunk_rows = limited_rows[start : start + max_rows_per_chunk]
        if not chunk_rows:
            continue
        row_start = chunk_rows[0][0]
        row_end = chunk_rows[-1][0]
        body = records_text(headers, ([rows[0]] if headers else []) + chunk_rows)
        if not body:
            continue
        text = (
            f"Spreadsheet file: {path.name}\n"
            f"Sheet: {sheet_name}\n"
            f"Rows: {row_start}-{row_end}\n"
            f"Headers: {header_text}\n\n"
            f"{body}"
        )
        chunks.append(
            Chunk(
                text,
                f"{sheet_name} rows {row_start}-{row_end}",
                source_type="spreadsheet",
                sheet=sheet_name,
                cell_range=f"{excel_col_name(first_col)}{row_start}:{excel_col_name(last_col)}{row_end}",
                row_start=row_start,
                row_end=row_end,
                column_headers=json.dumps(headers, ensure_ascii=False),
                has_formulas=bool(formulas),
                hidden_sheet=hidden,
            )
        )

    if len(rows) > row_limit:
        chunks.append(
            Chunk(
                (
                    f"Spreadsheet file: {path.name}\n"
                    f"Sheet: {sheet_name}\n"
                    f"Large sheet truncated for RAG indexing.\n"
                    f"Indexed rows: {row_limit}\n"
                    f"Total non-empty rows: {len(rows)}\n"
                    f"Headers: {header_text}"
                ),
                f"Large sheet digest: {sheet_name}",
                source_type="spreadsheet",
                sheet=sheet_name,
                cell_range=range_label,
                row_start=rows[0][0],
                row_end=last_row,
                column_headers=json.dumps(headers, ensure_ascii=False),
                has_formulas=bool(formulas),
                hidden_sheet=hidden,
            )
        )

    if env_bool("RAG_SPREADSHEET_INCLUDE_FORMULAS", True) and formulas:
        chunks.append(
            Chunk(
                f"Spreadsheet file: {path.name}\nSheet: {sheet_name}\nFormulas:\n" + "\n".join(formulas),
                f"Formulas: {sheet_name}",
                source_type="spreadsheet",
                sheet=sheet_name,
                cell_range=range_label,
                row_start=rows[0][0],
                row_end=last_row,
                column_headers=json.dumps(headers, ensure_ascii=False),
                has_formulas=True,
                hidden_sheet=hidden,
            )
        )

    if env_bool("RAG_SPREADSHEET_INCLUDE_COMMENTS", True) and comments:
        chunks.append(
            Chunk(
                f"Spreadsheet file: {path.name}\nSheet: {sheet_name}\nComments:\n" + "\n".join(comments),
                f"Comments: {sheet_name}",
                source_type="spreadsheet",
                sheet=sheet_name,
                cell_range=range_label,
                row_start=rows[0][0],
                row_end=last_row,
                column_headers=json.dumps(headers, ensure_ascii=False),
                has_formulas=bool(formulas),
                hidden_sheet=hidden,
            )
        )

    return chunks


def chunk_document(path: Path, text: str) -> tuple[dict[str, Any], list[Chunk]]:
    frontmatter, body = parse_frontmatter(text)
    title = extract_title(path, frontmatter, body)
    tags = extract_tags(body, frontmatter)
    links = extract_links(body)
    headings = [match.group(2).strip() for match in HEADING_RE.finditer(body)]
    if path.suffix.lower() == ".md":
        blocks = markdown_blocks(body)
    else:
        blocks = plain_blocks(body)
    chunks = trim_to_chunks(
        blocks,
        env_int("RAG_CHUNK_TOKENS", 800),
        env_int("RAG_CHUNK_OVERLAP_TOKENS", 120),
    )
    metadata = {
        "title": title,
        "frontmatter": frontmatter,
        "tags": tags,
        "links": links,
        "headings": headings,
    }
    return metadata, chunks


def ocr_enabled() -> bool:
    return env_bool("RAG_OCR_ENABLED", True) and env("RAG_OCR_MODE", "needed") in {"needed", "always"}


def ocr_languages() -> str:
    return env("RAG_OCR_LANGUAGES", "rus+eng+deu")


def ocr_tessdata_path() -> Path:
    raw = env("RAG_OCR_TESSDATA_PATH", ".runtime/tessdata")
    path = expand_path(raw)
    if not path.is_absolute():
        path = (PROJECT_ROOT / raw).resolve()
    return path


def ocr_tesseract_config() -> str:
    tessdata = ocr_tessdata_path()
    if tessdata.exists():
        return f"--tessdata-dir {shlex.quote(str(tessdata))}"
    return ""


def require_pymupdf() -> Any:
    try:
        import fitz
    except ImportError as exc:
        raise RuntimeError("pymupdf missing; run make rag-install") from exc
    return fitz


def require_ocr_modules() -> tuple[Any, Any]:
    try:
        from PIL import Image
        import pytesseract
    except ImportError as exc:
        raise RuntimeError("pillow/pytesseract missing; run make rag-install") from exc
    if not shutil.which("tesseract"):
        raise RuntimeError("ocr_required: tesseract binary missing; run make rag-install")
    return Image, pytesseract


def chunk_text_by_pages(
    path: Path,
    page_texts: list[tuple[int, str]],
    source_type: str,
    extractor: str,
    ocr: bool = False,
) -> tuple[dict[str, Any], list[Chunk]]:
    chunks: list[Chunk] = []
    headings = []
    languages = ocr_languages() if ocr else ""
    for page_num, text in page_texts:
        text = text.strip()
        if not text:
            continue
        page_heading = f"Page {page_num}" if source_type == "pdf" else "Image OCR"
        for index, chunk in enumerate(trim_to_chunks([(page_heading, text)], env_int("RAG_CHUNK_TOKENS", 800), env_int("RAG_CHUNK_OVERLAP_TOKENS", 120))):
            heading = page_heading if index == 0 else f"{page_heading} part {index + 1}"
            chunks.append(
                Chunk(
                    chunk.text,
                    heading,
                    source_type=source_type,
                    page=page_num,
                    page_start=page_num,
                    page_end=page_num,
                    ocr_used=ocr,
                    ocr_languages=languages,
                    extractor=extractor,
                )
            )
        headings.append(page_heading)
    return {
        "title": path.stem,
        "frontmatter": {
            "extractor": extractor,
            "ocr_used": ocr,
            "ocr_languages": languages,
            "page_count": len(page_texts),
        },
        "tags": [],
        "links": [],
        "headings": headings,
    }, chunks


def ocr_pil_image(image: Any) -> str:
    _, pytesseract = require_ocr_modules()
    return pytesseract.image_to_string(image, lang=ocr_languages(), config=ocr_tesseract_config()).strip()


def extract_pdf_ocr(path: Path, fitz: Any, doc: Any) -> tuple[dict[str, Any], list[Chunk]]:
    max_pages = min(len(doc), max(1, env_int("RAG_OCR_MAX_PAGES", 25)))
    dpi = max(72, env_int("RAG_OCR_DPI", 200))
    page_texts: list[tuple[int, str]] = []
    require_ocr_modules()
    for page_index in range(max_pages):
        page = doc[page_index]
        matrix = fitz.Matrix(dpi / 72.0, dpi / 72.0)
        pixmap = page.get_pixmap(matrix=matrix, alpha=False)
        from PIL import Image

        image = Image.open(io.BytesIO(pixmap.tobytes("png")))
        page_texts.append((page_index + 1, ocr_pil_image(image)))
    return chunk_text_by_pages(path, page_texts, "pdf", "tesseract", ocr=True)


def extract_pdf(path: Path) -> tuple[dict[str, Any], list[Chunk]]:
    if not env_bool("RAG_PDF_ENABLED", True):
        return {
            "title": path.stem,
            "frontmatter": {"extractor": "pdf-disabled"},
            "tags": [],
            "links": [],
            "headings": [],
        }, []
    fitz = require_pymupdf()
    doc = fitz.open(path)
    try:
        page_texts = [(index + 1, page.get_text("text").strip()) for index, page in enumerate(doc)]
        total_text = sum(len(text) for _, text in page_texts)
        if total_text >= env_int("RAG_OCR_MIN_TEXT_CHARS", 200):
            return chunk_text_by_pages(path, page_texts, "pdf", "pymupdf", ocr=False)
        if ocr_enabled():
            return extract_pdf_ocr(path, fitz, doc)
        raise RuntimeError("ocr_required: PDF text layer is empty or below RAG_OCR_MIN_TEXT_CHARS")
    finally:
        doc.close()


def extract_image(path: Path) -> tuple[dict[str, Any], list[Chunk]]:
    if not env_bool("RAG_IMAGES_ENABLED", True):
        return {
            "title": path.stem,
            "frontmatter": {"extractor": "images-disabled"},
            "tags": [],
            "links": [],
            "headings": [],
        }, []
    if not ocr_enabled():
        raise RuntimeError("ocr_required: image files require OCR")
    Image, _ = require_ocr_modules()
    with Image.open(path) as image:
        text = ocr_pil_image(image)
    return chunk_text_by_pages(path, [(1, text)], "image", "tesseract", ocr=True)


def extract_openpyxl_spreadsheet(path: Path) -> tuple[dict[str, Any], list[Chunk]]:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl missing; run make rag-install") from exc

    include_hidden = env_bool("RAG_SPREADSHEET_INCLUDE_HIDDEN", False)
    value_wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    formula_wb = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        chunks: list[Chunk] = []
        sheet_summaries = []

        for sheet_name in value_wb.sheetnames:
            value_ws = value_wb[sheet_name]
            formula_ws = formula_wb[sheet_name]
            hidden = formula_ws.sheet_state != "visible"
            if hidden and not include_hidden:
                continue

            raw_rows: list[tuple[int, list[str]]] = []
            for row in value_ws.iter_rows():
                raw_rows.append((row[0].row if row else len(raw_rows) + 1, [stringify_cell(cell.value) for cell in row]))
            rows, first_col = trim_sheet_rows(raw_rows)

            formulas: list[str] = []
            comments: list[str] = []
            for row in formula_ws.iter_rows():
                for cell in row:
                    value = cell.value
                    if isinstance(value, str) and value.startswith("="):
                        formulas.append(f"{cell.coordinate}: {value}")
                    if cell.comment and cell.comment.text:
                        comments.append(f"{cell.coordinate}: {cell.comment.text.strip()}")

            sheet_summaries.append(
                f"{sheet_name} ({'hidden' if hidden else 'visible'}, rows={max(0, len(rows) - 1)}, columns={max((len(row) for _, row in rows), default=0)})"
            )
            chunks.extend(
                spreadsheet_sheet_chunks(
                    path,
                    sheet_name=sheet_name,
                    rows=rows,
                    first_col=first_col,
                    hidden=hidden,
                    formulas=formulas,
                    comments=comments,
                )
            )

        defined_names = []
        try:
            for name in formula_wb.defined_names.values():
                defined_names.append(str(name))
        except Exception:
            defined_names = []

        metadata = {
            "title": path.stem,
            "frontmatter": {
                "extractor": "openpyxl",
                "workbook_sheets": value_wb.sheetnames,
                "defined_names": defined_names,
            },
            "tags": [],
            "links": [],
            "headings": [f"Workbook: {path.name}", *sheet_summaries],
        }
        if chunks:
            workbook_text = (
                f"Spreadsheet workbook: {path.name}\n"
                f"Extractor: openpyxl\n"
                f"Sheets:\n- " + "\n- ".join(sheet_summaries)
            )
            if defined_names:
                workbook_text += "\nNamed ranges:\n- " + "\n- ".join(defined_names)
            chunks.insert(0, Chunk(workbook_text, "Workbook summary", source_type="spreadsheet"))
        return metadata, chunks
    finally:
        value_wb.close()
        formula_wb.close()


def extract_calamine_spreadsheet(path: Path) -> tuple[dict[str, Any], list[Chunk]]:
    try:
        from python_calamine import CalamineWorkbook
    except ImportError as exc:
        raise RuntimeError("python-calamine missing; run make rag-install") from exc

    workbook = CalamineWorkbook.from_path(str(path))
    chunks: list[Chunk] = []
    sheet_summaries = []
    for sheet_name in workbook.sheet_names:
        sheet = workbook.get_sheet_by_name(sheet_name)
        raw_rows = [
            (index, [stringify_cell(value) for value in row])
            for index, row in enumerate(sheet.to_python(), start=1)
        ]
        rows, first_col = trim_sheet_rows(raw_rows)
        sheet_summaries.append(
            f"{sheet_name} (rows={max(0, len(rows) - 1)}, columns={max((len(row) for _, row in rows), default=0)})"
        )
        chunks.extend(
            spreadsheet_sheet_chunks(
                path,
                sheet_name=sheet_name,
                rows=rows,
                first_col=first_col,
            )
        )

    metadata = {
        "title": path.stem,
        "frontmatter": {
            "extractor": "python-calamine",
            "workbook_sheets": list(workbook.sheet_names),
        },
        "tags": [],
        "links": [],
        "headings": [f"Workbook: {path.name}", *sheet_summaries],
    }
    if chunks:
        chunks.insert(
            0,
            Chunk(
                (
                    f"Spreadsheet workbook: {path.name}\n"
                    f"Extractor: python-calamine\n"
                    f"Sheets:\n- " + "\n- ".join(sheet_summaries)
                ),
                "Workbook summary",
                source_type="spreadsheet",
            ),
        )
    return metadata, chunks


def extract_spreadsheet(path: Path) -> tuple[dict[str, Any], list[Chunk]]:
    if not env_bool("RAG_SPREADSHEETS_ENABLED", True):
        return {
            "title": path.stem,
            "frontmatter": {"extractor": "spreadsheet-disabled"},
            "tags": [],
            "links": [],
            "headings": [],
        }, []

    suffix = path.suffix.lower()
    errors = []
    if suffix in {".xlsx", ".xlsm"}:
        try:
            return extract_openpyxl_spreadsheet(path)
        except Exception as exc:
            errors.append(f"openpyxl: {exc}")

    try:
        return extract_calamine_spreadsheet(path)
    except Exception as exc:
        errors.append(f"python-calamine: {exc}")
        raise RuntimeError(f"spreadsheet extraction failed for {path.name}: {'; '.join(errors)}") from exc


def extract_document(path: Path) -> tuple[dict[str, Any], list[Chunk]]:
    suffix = path.suffix.lower()
    if suffix in SPREADSHEET_EXT_SET:
        return extract_spreadsheet(path)
    if suffix in PDF_EXT_SET:
        return extract_pdf(path)
    if suffix in IMAGE_EXT_SET:
        return extract_image(path)
    text = path.read_text(encoding="utf-8", errors="replace")
    return chunk_document(path, text)


class Embedder:
    def __init__(self) -> None:
        self.backend = env("RAG_EMBEDDING_BACKEND", "sentence-transformers")
        self.model_name = env("RAG_EMBEDDING_MODEL", "intfloat/multilingual-e5-small")
        self.model: Any = None
        if self.backend == "hash":
            return
        try:
            from sentence_transformers import SentenceTransformer
        except ImportError as exc:
            raise RuntimeError("sentence-transformers missing; run make rag-install") from exc
        self.model = SentenceTransformer(self.model_name)

    def embed(self, texts: list[str], query: bool = False) -> list[list[float]]:
        if self.backend == "hash":
            return [hash_embedding(text) for text in texts]
        prefix = "query: " if query else "passage: "
        vectors = self.model.encode(
            [prefix + text for text in texts],
            normalize_embeddings=True,
            show_progress_bar=False,
        )
        return [list(map(float, vector)) for vector in vectors]


def hash_embedding(text: str, dimensions: int = 384) -> list[float]:
    buckets = [0.0] * dimensions
    words = re.findall(r"[\w/-]+", text.lower())
    for word in words or [text.lower()]:
        digest = hashlib.sha256(word.encode("utf-8")).digest()
        index = int.from_bytes(digest[:4], "big") % dimensions
        sign = 1.0 if digest[4] % 2 == 0 else -1.0
        buckets[index] += sign
    norm = sum(value * value for value in buckets) ** 0.5 or 1.0
    return [value / norm for value in buckets]


def require_lancedb() -> Any:
    try:
        import lancedb
    except ImportError as exc:
        raise RuntimeError("lancedb missing; run make rag-install") from exc
    return lancedb


def db_table_names(db: Any) -> set[str]:
    if hasattr(db, "list_tables"):
        tables = db.list_tables()
        if hasattr(tables, "tables"):
            return set(tables.tables)
        return set(tables)
    return set(db.table_names())


def open_table(create: bool = False) -> Any:
    lancedb = require_lancedb()
    db = lancedb.connect(str(lancedb_path()))
    names = db_table_names(db)
    if TABLE_NAME in names:
        return db.open_table(TABLE_NAME)
    if create:
        return None
    raise RuntimeError("RAG index is empty; run make rag-index")


def drop_chunks_table() -> None:
    lancedb = require_lancedb()
    db = lancedb.connect(str(lancedb_path()))
    if TABLE_NAME in db_table_names(db):
        db.drop_table(TABLE_NAME)


def delete_document(table: Any, document_id: str) -> None:
    try:
        table.delete(f"document_id = '{document_id}'")
    except Exception:
        # Older lancedb versions throw when a predicate matches no rows.
        pass


def add_rows(rows: list[dict[str, Any]]) -> None:
    lancedb = require_lancedb()
    db = lancedb.connect(str(lancedb_path()))
    if TABLE_NAME in db_table_names(db):
        table = db.open_table(TABLE_NAME)
        table.add(rows)
    else:
        db.create_table(TABLE_NAME, data=rows)


def build_rows(root: Path, path: Path, digest: str, embedder: Embedder) -> list[dict[str, Any]]:
    metadata, chunks = extract_document(path)
    relative = relpath(path, root)
    stat = path.stat()
    vectors = embedder.embed([chunk.text for chunk in chunks]) if chunks else []
    rows: list[dict[str, Any]] = []
    document_id = hashlib.sha256(relative.encode("utf-8")).hexdigest()
    for index, chunk in enumerate(chunks):
        rows.append(
            {
                "id": f"{document_id}:{index}",
                "document_id": document_id,
                "relative_path": relative,
                "source_path": str(path),
                "chunk_index": index,
                "text": chunk.text,
                "title": metadata["title"],
                "heading": chunk.heading,
                "extension": path.suffix.lower(),
                "source_type": chunk.source_type,
                "sheet": chunk.sheet,
                "cell_range": chunk.cell_range,
                "row_start": int(chunk.row_start),
                "row_end": int(chunk.row_end),
                "column_headers": chunk.column_headers,
                "has_formulas": bool(chunk.has_formulas),
                "hidden_sheet": bool(chunk.hidden_sheet),
                "page": int(chunk.page),
                "page_start": int(chunk.page_start),
                "page_end": int(chunk.page_end),
                "ocr_used": bool(chunk.ocr_used),
                "ocr_languages": chunk.ocr_languages,
                "extractor": chunk.extractor,
                "mtime": float(stat.st_mtime),
                "sha256": digest,
                "tags": ",".join(metadata["tags"]),
                "links_json": json.dumps(metadata["links"], ensure_ascii=False),
                "frontmatter_json": json.dumps(metadata["frontmatter"], ensure_ascii=False),
                "headings_json": json.dumps(metadata["headings"], ensure_ascii=False),
                "vector": vectors[index],
            }
        )
    return rows


def index_documents(prune: bool = True) -> dict[str, Any]:
    root = source_path()
    if not root.exists():
        raise RuntimeError(f"RAG source path does not exist: {root}")
    index_path().mkdir(parents=True, exist_ok=True)
    lancedb_path().mkdir(parents=True, exist_ok=True)
    manifest = load_manifest()
    if manifest.get("table_schema_version") != TABLE_SCHEMA_VERSION:
        drop_chunks_table()
        manifest["documents"] = {}
        manifest["table_schema_version"] = TABLE_SCHEMA_VERSION
    documents = manifest.setdefault("documents", {})
    embedder = Embedder()
    files = discover_files(root)
    seen: set[str] = set()
    changed = 0
    skipped = 0
    chunks_written = 0
    table = open_table(create=True)

    for path in files:
        relative = relpath(path, root)
        seen.add(relative)
        digest = sha256_file(path)
        stat = path.stat()
        current = documents.get(relative, {})
        if current.get("sha256") == digest and not current.get("skipped"):
            skipped += 1
            continue
        document_id = hashlib.sha256(relative.encode("utf-8")).hexdigest()
        if table is not None:
            delete_document(table, document_id)
        extractor = ""
        try:
            rows = build_rows(root, path, digest, embedder)
            if rows:
                extractor = str(rows[0].get("extractor") or rows[0].get("source_type") or "text")
            elif path.suffix.lower() in SPREADSHEET_EXT_SET:
                extractor = "spreadsheet"
            elif path.suffix.lower() in PDF_EXT_SET:
                extractor = "pdf"
            elif path.suffix.lower() in IMAGE_EXT_SET:
                extractor = "image"
            else:
                extractor = "text"
        except Exception as exc:
            documents[relative] = {
                "document_id": document_id,
                "sha256": digest,
                "mtime": float(stat.st_mtime),
                "size": int(stat.st_size),
                "chunks": 0,
                "skipped": True,
                "error": str(exc),
                "extractor": "skipped",
            }
            changed += 1
            continue
        if rows:
            add_rows(rows)
            table = open_table(create=True)
        documents[relative] = {
            "document_id": document_id,
            "sha256": digest,
            "mtime": float(stat.st_mtime),
            "size": int(stat.st_size),
            "chunks": len(rows),
            "extractor": extractor,
        }
        changed += 1
        chunks_written += len(rows)

    pruned = 0
    if prune and table is not None:
        for relative in sorted(set(documents) - seen):
            delete_document(table, documents[relative]["document_id"])
            del documents[relative]
            pruned += 1

    manifest["updated_at"] = time.time()
    manifest["source_path"] = str(root)
    manifest["embedding_model"] = env("RAG_EMBEDDING_MODEL", "intfloat/multilingual-e5-small")
    manifest["embedding_backend"] = env("RAG_EMBEDDING_BACKEND", "sentence-transformers")
    save_manifest(manifest)
    return {
        "source_path": str(root),
        "documents": len(documents),
        "changed": changed,
        "skipped": skipped,
        "pruned": pruned,
        "chunks_written": chunks_written,
    }


def row_matches(row: dict[str, Any], filters: dict[str, Any]) -> bool:
    path_filter = filters.get("path")
    if path_filter and path_filter not in row.get("relative_path", ""):
        return False
    ext_filter = filters.get("extension")
    if ext_filter:
        ext = str(ext_filter)
        if not ext.startswith("."):
            ext = "." + ext
        if row.get("extension") != ext:
            return False
    tag_filter = filters.get("tag")
    if tag_filter:
        tags = {tag.strip() for tag in str(row.get("tags", "")).split(",") if tag.strip()}
        if str(tag_filter).lstrip("#") not in tags:
            return False
    source_type_filter = filters.get("source_type")
    if source_type_filter and row.get("source_type") != source_type_filter:
        return False
    page_filter = filters.get("page")
    if page_filter:
        try:
            if int(row.get("page", 0)) != int(page_filter):
                return False
        except (TypeError, ValueError):
            return False
    sheet_filter = filters.get("sheet")
    if sheet_filter and str(sheet_filter).lower() not in str(row.get("sheet", "")).lower():
        return False
    modified_after = filters.get("modified_after")
    if modified_after:
        try:
            if float(row.get("mtime", 0)) < float(modified_after):
                return False
        except ValueError:
            return False
    return True


def search(query: str, top_k: int | None = None, filters: dict[str, Any] | None = None) -> dict[str, Any]:
    filters = filters or {}
    top_k = top_k or env_int("RAG_TOP_K", 8)
    table = open_table()
    embedder = Embedder()
    vector = embedder.embed([query], query=True)[0]
    # Fetch a wider candidate set so CLI-side filters still leave enough rows.
    rows = table.search(vector).limit(max(top_k * 5, top_k)).to_list()
    results = []
    for row in rows:
        if not row_matches(row, filters):
            continue
        text = str(row.get("text", ""))
        results.append(
            {
                "id": row.get("id"),
                "document_id": row.get("document_id"),
                "path": row.get("relative_path"),
                "title": row.get("title"),
                "heading": row.get("heading"),
                "extension": row.get("extension"),
                "source_type": row.get("source_type"),
                "sheet": row.get("sheet"),
                "range": row.get("cell_range"),
                "row_start": row.get("row_start"),
                "row_end": row.get("row_end"),
                "column_headers": row.get("column_headers"),
                "has_formulas": row.get("has_formulas"),
                "hidden_sheet": row.get("hidden_sheet"),
                "page": row.get("page"),
                "page_start": row.get("page_start"),
                "page_end": row.get("page_end"),
                "ocr_used": row.get("ocr_used"),
                "ocr_languages": row.get("ocr_languages"),
                "extractor": row.get("extractor"),
                "score": float(row.get("_distance", 0.0)),
                "tags": [tag for tag in str(row.get("tags", "")).split(",") if tag],
                "excerpt": make_excerpt(text),
                "text": text,
            }
        )
        if len(results) >= top_k:
            break
    return {"query": query, "top_k": top_k, "results": results}


def make_excerpt(text: str, max_chars: int = 500) -> str:
    text = re.sub(r"\s+", " ", text).strip()
    if len(text) <= max_chars:
        return text
    return text[: max_chars - 1].rstrip() + "..."


def render_markdown(payload: dict[str, Any]) -> str:
    results = payload.get("results", [])
    if not results:
        return f"No RAG results for: {payload.get('query', '')}"
    lines = [f"RAG results for: {payload.get('query', '')}", ""]
    for idx, item in enumerate(results, start=1):
        heading = item.get("heading") or item.get("title") or item.get("path")
        lines.append(f"{idx}. {item.get('path')} - {heading}")
        details = []
        if item.get("sheet"):
            details.append(f"sheet: {item.get('sheet')}")
        if item.get("range"):
            details.append(f"range: {item.get('range')}")
        if item.get("page"):
            details.append(f"page: {item.get('page')}")
        if item.get("ocr_used"):
            details.append("ocr: yes")
        if item.get("extractor"):
            details.append(f"extractor: {item.get('extractor')}")
        if item.get("source_type"):
            details.append(f"type: {item.get('source_type')}")
        if details:
            lines.append(f"   {'; '.join(details)}")
        lines.append(f"   score: {item.get('score'):.4f}")
        if item.get("tags"):
            lines.append(f"   tags: {', '.join(item['tags'])}")
        lines.append(f"   {item.get('excerpt', '')}")
        lines.append("")
    return "\n".join(lines).rstrip()


def get_document(document_id: str) -> dict[str, Any]:
    table = open_table()
    rows = table.search().where(f"document_id = '{document_id}'").limit(1000).to_list()
    rows = sorted(rows, key=lambda row: int(row.get("chunk_index", 0)))
    if not rows:
        raise KeyError(document_id)
    return {
        "document_id": document_id,
        "path": rows[0].get("relative_path"),
        "title": rows[0].get("title"),
        "chunks": [
            {
                "chunk_index": row.get("chunk_index"),
                "heading": row.get("heading"),
                "source_type": row.get("source_type"),
                "sheet": row.get("sheet"),
                "range": row.get("cell_range"),
                "page": row.get("page"),
                "page_start": row.get("page_start"),
                "page_end": row.get("page_end"),
                "ocr_used": row.get("ocr_used"),
                "ocr_languages": row.get("ocr_languages"),
                "extractor": row.get("extractor"),
                "text": row.get("text"),
            }
            for row in rows
        ],
    }


def health() -> dict[str, Any]:
    manifest = load_manifest()
    table_exists = False
    try:
        table_exists = TABLE_NAME in db_table_names(require_lancedb().connect(str(lancedb_path())))
    except Exception:
        table_exists = False
    return {
        "ok": True,
        "enabled": env_bool("RAG_ENABLED", True),
        "source_path": env("RAG_SOURCE_PATH", env("OBSIDIAN_SHARED_PATH", "")),
        "index_path": str(index_path()),
        "table_exists": table_exists,
        "documents": len(manifest.get("documents", {})),
        "table_schema_version": manifest.get("table_schema_version"),
        "spreadsheets_enabled": env_bool("RAG_SPREADSHEETS_ENABLED", True),
        "pdf_enabled": env_bool("RAG_PDF_ENABLED", True),
        "images_enabled": env_bool("RAG_IMAGES_ENABLED", True),
        "ocr_enabled": env_bool("RAG_OCR_ENABLED", True),
        "ocr_mode": env("RAG_OCR_MODE", "needed"),
        "ocr_languages": ocr_languages(),
        "embedding_model": env("RAG_EMBEDDING_MODEL", "intfloat/multilingual-e5-small"),
        "embedding_backend": env("RAG_EMBEDDING_BACKEND", "sentence-transformers"),
    }


def source_fingerprint(root: Path) -> str:
    digest = hashlib.sha256()
    for path in discover_files(root):
        try:
            stat = path.stat()
        except OSError:
            continue
        relative = relpath(path, root)
        digest.update(relative.encode("utf-8", errors="replace"))
        digest.update(b"\0")
        digest.update(str(stat.st_mtime_ns).encode("ascii"))
        digest.update(b"\0")
        digest.update(str(stat.st_size).encode("ascii"))
        digest.update(b"\0")
    return digest.hexdigest()


def watch(interval: int | None = None, debounce: int | None = None) -> None:
    interval = interval or env_int("RAG_WATCH_INTERVAL_SECONDS", 20)
    debounce = debounce if debounce is not None else env_int("RAG_WATCH_DEBOUNCE_SECONDS", 3)
    interval = max(2, interval)
    debounce = max(0, debounce)
    root = source_path()
    previous = ""
    print(
        json.dumps(
            {
                "event": "rag-watch-start",
                "source_path": str(root),
                "index_path": str(index_path()),
                "interval_seconds": interval,
                "debounce_seconds": debounce,
            },
            ensure_ascii=False,
        ),
        flush=True,
    )
    while True:
        try:
            current = source_fingerprint(root)
            if current != previous:
                if previous and debounce:
                    time.sleep(debounce)
                    current = source_fingerprint(root)
                result = index_documents(prune=True)
                result["event"] = "rag-index"
                print(json.dumps(result, ensure_ascii=False, sort_keys=True), flush=True)
                previous = current
        except Exception as exc:
            print(json.dumps({"event": "rag-watch-error", "error": str(exc)}, ensure_ascii=False), flush=True)
        time.sleep(interval)


def serve() -> None:
    try:
        from fastapi import FastAPI, HTTPException, Request
        import uvicorn
    except ImportError as exc:
        raise RuntimeError("FastAPI/uvicorn missing; run make rag-install") from exc

    # With postponed annotations enabled, FastAPI resolves route annotations from
    # module globals rather than this function's local imports.
    globals()["Request"] = Request

    app = FastAPI(title="oMLX Agent Local RAG", version="0.5.2")

    @app.get("/health")
    def health_route() -> dict[str, Any]:
        return health()

    @app.post("/index")
    async def index_route(request: Request) -> dict[str, Any]:
        try:
            payload = await request.json()
        except Exception:
            payload = {}
        prune = bool(payload.get("prune", True)) if isinstance(payload, dict) else True
        return index_documents(prune=prune)

    @app.post("/search")
    async def search_route(request: Request) -> dict[str, Any]:
        try:
            payload = await request.json()
        except Exception as exc:
            raise HTTPException(status_code=400, detail="invalid JSON body") from exc
        if not isinstance(payload, dict):
            raise HTTPException(status_code=400, detail="JSON body must be an object")
        query = str(payload.get("query", "")).strip()
        if not query:
            raise HTTPException(status_code=400, detail="query is required")
        top_k = payload.get("top_k")
        if top_k is not None:
            try:
                top_k = int(top_k)
            except (TypeError, ValueError) as exc:
                raise HTTPException(status_code=400, detail="top_k must be an integer") from exc
            if top_k < 1 or top_k > 50:
                raise HTTPException(status_code=400, detail="top_k must be between 1 and 50")
        filters = payload.get("filters") or {}
        if not isinstance(filters, dict):
            raise HTTPException(status_code=400, detail="filters must be an object")
        return search(query, top_k=top_k, filters=filters)

    @app.get("/documents/{document_id}")
    def document_route(document_id: str) -> dict[str, Any]:
        try:
            return get_document(document_id)
        except KeyError as exc:
            raise HTTPException(status_code=404, detail="document not found") from exc

    host = env("RAG_BIND_HOST", env("RAG_HOST", "127.0.0.1"))
    port = env_int("RAG_PORT", 8765)
    uvicorn.run(app, host=host, port=port)


def doctor() -> int:
    ok = True
    print(f"enabled={env('RAG_ENABLED', '1')}")
    raw_source = env("RAG_SOURCE_PATH", env("OBSIDIAN_SHARED_PATH", ""))
    if raw_source in {"", "${OBSIDIAN_SHARED_PATH}", "${OBSIDIAN_SHARED_PATH:-}"}:
        raw_source = env("OBSIDIAN_SHARED_PATH", "")
    if raw_source:
        src = expand_path(raw_source)
        print(f"source={src}")
        if src.exists() and src.is_dir():
            print("source_status=ok")
        else:
            print("source_status=missing")
            ok = False
    else:
        print("source_status=unset")
        ok = False
    print(f"index={index_path()}")
    for module in ("lancedb", "fastapi", "uvicorn"):
        try:
            __import__(module)
            print(f"module_{module}=ok")
        except ImportError:
            print(f"module_{module}=missing")
            ok = False
    if env_bool("RAG_SPREADSHEETS_ENABLED", True):
        for module in ("openpyxl", "python_calamine"):
            try:
                __import__(module)
                print(f"module_{module}=ok")
            except ImportError:
                print(f"module_{module}=missing")
                ok = False
        try:
            __import__("duckdb")
            print("module_duckdb=ok")
        except ImportError:
            print("module_duckdb=missing optional")
    if env_bool("RAG_PDF_ENABLED", True):
        try:
            __import__("fitz")
            print("module_pymupdf=ok")
        except ImportError:
            print("module_pymupdf=missing")
            ok = False
    if env_bool("RAG_IMAGES_ENABLED", True) or env_bool("RAG_OCR_ENABLED", True):
        for module in ("PIL", "pytesseract"):
            try:
                __import__(module)
                print(f"module_{module}=ok")
            except ImportError:
                print(f"module_{module}=missing")
                ok = False
    if env_bool("RAG_OCR_ENABLED", True):
        tesseract_path = shutil.which("tesseract")
        local_tessdata = ocr_tessdata_path()
        print(f"ocr_tessdata={local_tessdata}")
        if tesseract_path:
            print(f"binary_tesseract=ok path={tesseract_path}")
            try:
                cmd = ["tesseract", "--list-langs"]
                if local_tessdata.exists():
                    cmd.extend(["--tessdata-dir", str(local_tessdata)])
                output = subprocess.check_output(cmd, text=True, stderr=subprocess.STDOUT)
                langs = sorted(line.strip() for line in output.splitlines()[1:] if line.strip())
                print(f"ocr_languages_installed={','.join(langs)}")
                wanted = set(ocr_languages().split("+"))
                local_langs = {path.stem for path in local_tessdata.glob("*.traineddata")} if local_tessdata.exists() else set()
                missing = sorted(wanted - set(langs) - local_langs)
                if local_langs:
                    print(f"ocr_languages_local={','.join(sorted(local_langs))}")
                if missing:
                    print(f"ocr_languages_missing={','.join(missing)}")
                    ok = False
            except Exception as exc:
                print(f"ocr_languages_error={exc}")
                ok = False
        else:
            print("binary_tesseract=missing")
            ok = False
    if env("RAG_EMBEDDING_BACKEND", "sentence-transformers") != "hash":
        try:
            __import__("sentence_transformers")
            print("module_sentence_transformers=ok")
        except ImportError:
            print("module_sentence_transformers=missing")
            ok = False
    print(json.dumps(health(), sort_keys=True))
    return 0 if ok else 2


def self_test() -> None:
    class ListTablesObject:
        tables = ["chunks", "archive"]

    class NewLanceDb:
        def list_tables(self) -> ListTablesObject:
            return ListTablesObject()

    class ListLanceDb:
        def list_tables(self) -> list[str]:
            return ["chunks"]

    class OldLanceDb:
        def table_names(self) -> list[str]:
            return ["legacy"]

    assert db_table_names(NewLanceDb()) == {"chunks", "archive"}
    assert db_table_names(ListLanceDb()) == {"chunks"}
    assert db_table_names(OldLanceDb()) == {"legacy"}

    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp) / "vault"
        root.mkdir()
        (root / ".obsidian").mkdir()
        (root / ".obsidian" / "workspace.json").write_text('{"secret": true}', encoding="utf-8")
        (root / "Nested" / "Vault" / ".obsidian").mkdir(parents=True)
        (root / "Nested" / "Vault" / ".obsidian" / "plugins.json").write_text('{"secret": true}', encoding="utf-8")
        (root / "Nested" / "Vault" / "nested.md").write_text("# Nested\n\nVisible note.", encoding="utf-8")
        (root / "note.md").write_text(
            "---\ntitle: Local RAG\ntags: [agent, obsidian]\n---\n"
            "# Local RAG\n\nThis note explains [[Hermes]] and OpenClaw search.\n\n# Details\n\nMore text.",
            encoding="utf-8",
        )
        (root / "data.json").write_text('{"project": "omlx", "kind": "rag"}', encoding="utf-8")
        (root / "secret.env").write_text("TOKEN=bad", encoding="utf-8")
        previous = dict(os.environ)
        try:
            os.environ["RAG_SOURCE_PATH"] = str(root)
            os.environ["RAG_TEXT_EXTENSIONS"] = ".md,.json"
            os.environ["RAG_EXCLUDE_GLOBS"] = DEFAULT_EXCLUDES
            files = [relpath(path, root) for path in discover_files(root)]
            assert set(files) == {"Nested/Vault/nested.md", "data.json", "note.md"}, files
            metadata, chunks = chunk_document(root / "note.md", (root / "note.md").read_text())
            assert metadata["title"] == "Local RAG"
            assert "agent" in metadata["tags"]
            assert "Hermes" in metadata["links"]
            assert chunks and "OpenClaw" in " ".join(chunk.text for chunk in chunks)
            sheet_rows = [
                (1, ["Project", "Budget", "Status"]),
                (2, ["Hermes", "1200", "Active"]),
                (3, ["OpenClaw", "900", "Pilot"]),
            ]
            sheet_chunks = spreadsheet_sheet_chunks(root / "budget.xlsx", "Budget", sheet_rows, 1)
            assert any("OpenClaw" in chunk.text and chunk.sheet == "Budget" for chunk in sheet_chunks)
            assert any(chunk.cell_range == "A2:C3" for chunk in sheet_chunks), [chunk.cell_range for chunk in sheet_chunks]
        finally:
            os.environ.clear()
            os.environ.update(previous)

    try:
        import openpyxl
    except ImportError:
        pass
    else:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "rag-budget.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Budget"
            ws.append(["Project", "Budget", "Status"])
            ws.append(["Hermes", 1200, "Active"])
            ws.append(["OpenClaw", 900, "Pilot"])
            ws["D1"] = "Total"
            ws["D2"] = "=SUM(B2:B3)"
            ws["A2"].comment = openpyxl.comments.Comment("Core agent stack", "RAG")
            hidden = wb.create_sheet("Hidden")
            hidden.sheet_state = "hidden"
            hidden.append(["Secret"])
            hidden.append(["Should not index"])
            wb.save(path)
            wb.close()
            previous = dict(os.environ)
            try:
                os.environ["RAG_SPREADSHEET_INCLUDE_HIDDEN"] = "0"
                metadata, chunks = extract_spreadsheet(path)
                joined = "\n".join(chunk.text for chunk in chunks)
                assert metadata["frontmatter"]["extractor"] == "openpyxl"
                assert "OpenClaw" in joined
                assert "=SUM(B2:B3)" in joined
                assert "Core agent stack" in joined
                assert "Should not index" not in joined
            finally:
                os.environ.clear()
                os.environ.update(previous)
    try:
        import fitz
    except ImportError:
        pass
    else:
        with tempfile.TemporaryDirectory() as tmp:
            pdf_path = Path(tmp) / "rag-text.pdf"
            doc = fitz.open()
            page = doc.new_page()
            page.insert_text((72, 72), "pdf-text-sentinel from selectable text PDF")
            doc.save(pdf_path)
            doc.close()
            previous = dict(os.environ)
            try:
                os.environ["RAG_OCR_ENABLED"] = "0"
                os.environ["RAG_OCR_MIN_TEXT_CHARS"] = "10"
                metadata, chunks = extract_pdf(pdf_path)
                joined = "\n".join(chunk.text for chunk in chunks)
                assert metadata["frontmatter"]["extractor"] == "pymupdf"
                assert "pdf-text-sentinel" in joined
                assert any(chunk.source_type == "pdf" and chunk.page == 1 and not chunk.ocr_used for chunk in chunks)
            finally:
                os.environ.clear()
                os.environ.update(previous)
    try:
        from PIL import Image, ImageDraw
    except ImportError:
        pass
    else:
        with tempfile.TemporaryDirectory() as tmp:
            image_path = Path(tmp) / "scan.png"
            image = Image.new("RGB", (500, 120), "white")
            draw = ImageDraw.Draw(image)
            draw.text((20, 40), "ocr-required-sentinel", fill="black")
            image.save(image_path)
            previous = dict(os.environ)
            try:
                os.environ["RAG_OCR_ENABLED"] = "0"
                try:
                    extract_image(image_path)
                    raise AssertionError("image extraction should require OCR when OCR is disabled")
                except RuntimeError as exc:
                    assert "ocr_required" in str(exc)
            finally:
                os.environ.clear()
                os.environ.update(previous)
    print("rag self-test ok")


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Local LanceDB RAG for Obsidian/text files.")
    sub = parser.add_subparsers(dest="command", required=True)
    sub.add_parser("doctor")
    sub.add_parser("index").add_argument("--no-prune", action="store_true")
    search_parser = sub.add_parser("search")
    search_parser.add_argument("query", nargs="*")
    search_parser.add_argument("--query", dest="query_opt")
    search_parser.add_argument("--top-k", type=int)
    search_parser.add_argument("--json", action="store_true")
    search_parser.add_argument("--path")
    search_parser.add_argument("--extension")
    search_parser.add_argument("--source-type")
    search_parser.add_argument("--page")
    search_parser.add_argument("--sheet")
    search_parser.add_argument("--tag")
    search_parser.add_argument("--modified-after")
    sub.add_parser("serve")
    watch_parser = sub.add_parser("watch")
    watch_parser.add_argument("--interval", type=int)
    watch_parser.add_argument("--debounce", type=int)
    sub.add_parser("health")
    sub.add_parser("self-test")
    args = parser.parse_args(argv)

    try:
        if args.command == "doctor":
            return doctor()
        if args.command == "index":
            print(json.dumps(index_documents(prune=not args.no_prune), ensure_ascii=False, indent=2))
            return 0
        if args.command == "search":
            query = args.query_opt or " ".join(args.query).strip()
            if not query:
                raise RuntimeError("query is required")
            filters = {
                key: value
                for key, value in {
                    "path": args.path,
                    "extension": args.extension,
                    "source_type": args.source_type,
                    "page": args.page,
                    "sheet": args.sheet,
                    "tag": args.tag,
                    "modified_after": args.modified_after,
                }.items()
                if value
            }
            payload = search(query, top_k=args.top_k, filters=filters)
            if args.json:
                print(json.dumps(payload, ensure_ascii=False, indent=2))
            else:
                print(render_markdown(payload))
            return 0
        if args.command == "serve":
            serve()
            return 0
        if args.command == "watch":
            watch(interval=args.interval, debounce=args.debounce)
            return 0
        if args.command == "health":
            print(json.dumps(health(), ensure_ascii=False, indent=2, sort_keys=True))
            return 0
        if args.command == "self-test":
            self_test()
            return 0
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1
    return 1


if __name__ == "__main__":
    raise SystemExit(main())
